# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import tempfile
from io import BytesIO
from typing import Any, Dict, Iterable, Optional, Set, Tuple

import openpyxl
import pandas as pd
import streamlit as st
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from sqlalchemy.exc import SQLAlchemyError

# ============================
# Config da página
# ============================
st.set_page_config(
    page_title="Agente de Estoque",
    page_icon="📦",
    layout="wide",
)

# ============================
# Session State
# ============================
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame()
if "comandos_acumulados" not in st.session_state:
    st.session_state.comandos_acumulados: list[tuple[str, str]] = []
if "df_source" not in st.session_state:
    st.session_state.df_source: Optional[str] = None  # 'db' ou 'upload'

# ============================
# Constantes
# ============================
TABLE_NAME = "estoque_end"
SCHEMA_NAME = "public"

# ============================
# Helpers de dados
# ============================
def _normalize_blanks(df: pd.DataFrame) -> pd.DataFrame:
    """Converte strings de espaços puros em '', e NaN para ''."""
    if df.empty:
        return df
    return df.replace(r"^\s+$", "", regex=True).fillna("")


def aplicar_regra_zero_vazio(df: pd.DataFrame) -> pd.DataFrame:
    """Se QTD == 0 ⇒ SKU = 'VAZIO' e DESCRIÇÃO = 'VAZIO'. Garante QTD >= 0."""
    if df.empty:
        return df
    df = df.copy()
    if "QTD" not in df.columns:
        return df
    df["QTD"] = pd.to_numeric(df["QTD"], errors="coerce").fillna(0).astype(int).clip(lower=0)
    zero_mask = df["QTD"].eq(0)
    if "SKU" in df.columns:
        df.loc[zero_mask, "SKU"] = "VAZIO"
    if "DESCRIÇÃO" in df.columns:
        df.loc[zero_mask, "DESCRIÇÃO"] = "VAZIO"
    return df


def _resolve_db_password() -> str:
    """Busca senha do DB em st.secrets['DB_PASSWORD'] ou env DB_PASSWORD."""
    try:
        pw = st.secrets.get("DB_PASSWORD")  # type: ignore[attr-defined]
        if pw:
            return str(pw)
    except Exception:
        pass
    env_pw = os.getenv("DB_PASSWORD")
    if env_pw:
        return env_pw
    raise ValueError(
        "A senha do banco não foi encontrada. Defina `st.secrets['DB_PASSWORD']` "
        "ou a variável de ambiente 'DB_PASSWORD'."
    )


@st.cache_resource(show_spinner=False)
def _get_engine() -> Engine:
    """Constroi e cacheia o engine do Postgres."""
    db_config = {
        "dbname": "OperacaoVIVO",
        "user": "postgres",
        "password": _resolve_db_password(),
        "host": "localhost",
        "port": 5432,
    }
    url = (
        f"postgresql+psycopg2://{db_config['user']}:{db_config['password']}"
        f"@{db_config['host']}:{db_config['port']}/{db_config['dbname']}"
    )
    return create_engine(url, echo=False, pool_pre_ping=True, future=True)


def atualizar_endereco(df: pd.DataFrame) -> pd.DataFrame:
    """Padroniza FILA, TORRE, NÍVEL (upper/strip) e recalcula ENDEREÇO = FILA-TORRE-NÍVEL."""
    if df.empty:
        return df
    df = df.copy()
    required = ["FILA", "TORRE", "NÍVEL"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes: {missing}")
    for col in required:
        df[col] = df[col].astype(str).str.upper().str.strip()
    df["ENDEREÇO"] = df[["FILA", "TORRE", "NÍVEL"]].agg("-".join, axis=1)
    return df


def _load_excel_to_df(uploaded_file: Any) -> pd.DataFrame:
    """Lê Excel (sheet 0), normaliza vazios, recalcula ENDEREÇO e aplica regra QTD==0."""
    if uploaded_file is None:
        return pd.DataFrame()
    try:
        uploaded_file.seek(0)
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, engine="openpyxl", dtype=str)
    except Exception:
        tmp_path: Optional[str] = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(uploaded_file.getbuffer())
                tmp_path = tmp.name
            df_raw = pd.read_excel(tmp_path, sheet_name=0, engine="openpyxl", dtype=str)
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
    df = _normalize_blanks(df_raw)
    df = atualizar_endereco(df)
    df = aplicar_regra_zero_vazio(df)
    return df


def _load_db_to_df() -> pd.DataFrame:
    """Lê do Postgres, normaliza vazios, recalcula ENDEREÇO e aplica regra QTD==0."""
    try:
        engine = _get_engine()
        with engine.connect() as conn:
            df = pd.read_sql(text(f"SELECT * FROM {SCHEMA_NAME}.{TABLE_NAME}"), conn)
        df = _normalize_blanks(df)
        df = atualizar_endereco(df)
        df = aplicar_regra_zero_vazio(df)
        return df
    except Exception as e:
        st.error(f"Erro ao carregar dados do banco de dados: {e}")
        return pd.DataFrame()


def save_to_db(df: pd.DataFrame) -> None:
    """Salva o DataFrame em public.estoque_end (replace, transacional)."""
    if df is None or df.empty:
        st.warning("Não há dados para salvar.")
        return
    try:
        engine = _get_engine()
        df_to_save = aplicar_regra_zero_vazio(df.copy())
        with engine.begin() as conn:
            df_to_save.to_sql(
                TABLE_NAME,
                con=conn,
                if_exists="replace",
                index=False,
                method="multi",
                schema=SCHEMA_NAME,
            )
        st.success("Dados registrados com sucesso no banco de dados.")
    except SQLAlchemyError as e:
        st.error(f"Erro ao salvar dados no banco de dados: {e}")


def drop_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Remove duplicidades por ENDEREÇO + SKU (fallback: todas)."""
    if df.empty:
        return df
    subset_cols = [c for c in ["ENDEREÇO", "SKU"] if c in df.columns]
    if subset_cols:
        return df.drop_duplicates(subset=subset_cols).reset_index(drop=True)
    return df.drop_duplicates().reset_index(drop=True)


def regra_sku_vazio(df: pd.DataFrame, enderecos_afetados: Optional[Iterable[str]] = None) -> pd.DataFrame:
    """Em cada ENDEREÇO com algum SKU real, remove linhas cujo SKU é VAZIO/''."""
    if df.empty:
        return df
    df = df.copy()
    if "SKU" not in df.columns or "ENDEREÇO" not in df.columns:
        return df
    sku_norm = df["SKU"].fillna("").astype(str).str.strip().str.upper()
    end_norm = df["ENDEREÇO"].fillna("").astype(str).str.strip().str.upper()
    is_vazio_logico = sku_norm.isin(["", "VAZIO"])
    has_non_vazio = (~is_vazio_logico).groupby(end_norm, dropna=False).transform("any")
    drop_mask = is_vazio_logico & has_non_vazio
    if enderecos_afetados:
        alvo = {str(e).strip().upper() for e in enderecos_afetados}
        drop_mask = drop_mask & end_norm.isin(alvo)
    return df.loc[~drop_mask].reset_index(drop=True)

# ============================
# Helpers de comandos
# ============================
# Regex pré-compiladas (aceitam 'endereço' e 'endereco')
P_MUDAR = re.compile(
    r"""^\s*mudar\s+sku\s+(?P<sku>\S+)\s+(?:do|no)\s+endere(?:ç|c)o\s+
        (?P<origem>\S+)\s+para\s+(?:o|a)\s+(?P<destino>\S+)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)
P_TROCAR = re.compile(
    r"""^\s*trocar\s+sku\s+(?P<sku1>\S+)\s+(?:do|no)\s+endere(?:ç|c)o\s+(?P<end1>\S+)\s+
        com\s+o\s+(?P<sku2>\S+)\s+(?:do|no)\s+endere(?:ç|c)o\s+(?P<end2>\S+)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)
P_TIRAR = re.compile(
    r"""^\s*tirar\s+(?P<qtd>\d+)\s+(?P<sku>\S+)\s+(?:do|no)\s+endere(?:ç|c)o\s+(?P<end>\S+)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)
P_ADD = re.compile(
    r"""^\s*add\s+(?P<qtd>\d+)\s+(?P<sku>\S+)\s+no\s+endere(?:ç|c)o\s+(?P<end>\S+)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)
P_LIMPAR = re.compile(
    r"""^\s*limpar\s+endere(?:ç|c)o\s+(?P<end>\S+)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)

def _norm(s: object) -> str:
    return str(s).upper().strip()

def _split_endereco(end: str) -> tuple[str, str, str]:
    parts = _norm(end).split("-")
    if len(parts) < 2:
        raise ValueError("Endereço inválido. Use o formato FILA-TORRE-NÍVEL (nível opcional).")
    fila, torre = parts[0], parts[1]
    nivel = parts[2] if len(parts) > 2 and parts[2] else ""
    return fila, torre, nivel

def _mask_sku_endereco(df: pd.DataFrame, sku: str, end: str) -> pd.Series:
    return (df["SKU"].astype(str).str.upper().str.strip() == _norm(sku)) & (
        df["ENDEREÇO"].astype(str).str.upper().str.strip() == _norm(end)
    )

# ============================
# Ações dos comandos
# ============================
def _cmd_mudar(df: pd.DataFrame, sku: str, origem: str, destino: str) -> tuple[pd.DataFrame, str, set[str]]:
    df = df.copy()
    sku_u, origem_u, destino_u = _norm(sku), _norm(origem), _norm(destino)
    enderecos_tocados: set[str] = {origem_u, destino_u}

    try:
        fila_d, torre_d, nivel_d = _split_endereco(destino_u)
    except ValueError as e:
        return df, f"{e}", enderecos_tocados

    cond_origem = _mask_sku_endereco(df, sku_u, origem_u)
    if not cond_origem.any():
        return df, f"SKU {sku} no endereço {origem_u} não encontrado.", enderecos_tocados

    qtd_move = pd.to_numeric(df.loc[cond_origem, "QTD"], errors="coerce").fillna(0).astype(int).sum()
    desc_ref_series = df.loc[cond_origem, "DESCRIÇÃO"].astype(str).str.strip()
    desc_ref = next((d for d in desc_ref_series if d), "")

    if qtd_move <= 0:
        df.loc[cond_origem, ["QTD", "SKU", "DESCRIÇÃO"]] = [0, "VAZIO", "VAZIO"]
        df = atualizar_endereco(df)
        return df, f"Não há quantidade para mover. Origem {origem_u} marcada como VAZIO.", enderecos_tocados

    cond_dest = df["ENDEREÇO"].astype(str).str.upper().str.strip() == destino_u
    sku_upper = df["SKU"].astype(str).str.upper().str.strip()
    cond_dest_vazio = cond_dest & (sku_upper == "VAZIO")
    cond_dest_same = cond_dest & (sku_upper == sku_u)
    cond_dest_occupied = cond_dest & (~sku_upper.isin(["", "VAZIO", sku_u]))

    if cond_dest_occupied.any():
        return (
            df,
            f"Endereço {destino_u} já ocupado por outro SKU. Use 'trocar' ou 'limpar endereço {destino_u}' antes de mover.",
            enderecos_tocados,
        )

    if cond_dest_same.any():
        df.loc[cond_dest_same, "QTD"] = (
            pd.to_numeric(df.loc[cond_dest_same, "QTD"], errors="coerce").fillna(0).astype(int) + int(qtd_move)
        )
        df.loc[cond_dest_same, ["FILA", "TORRE", "NÍVEL"]] = [fila_d, torre_d, nivel_d]
    elif cond_dest_vazio.any():
        df.loc[cond_dest_vazio, ["SKU", "DESCRIÇÃO", "FILA", "TORRE", "NÍVEL", "QTD"]] = [
            sku_u,
            desc_ref,
            fila_d,
            torre_d,
            nivel_d,
            int(qtd_move),
        ]
    else:
        nova_linha = {
            "SKU": sku_u,
            "DESCRIÇÃO": desc_ref,
            "FILA": fila_d,
            "TORRE": torre_d,
            "NÍVEL": nivel_d,
            "ENDEREÇO": destino_u,
            "QTD": int(qtd_move),
        }
        df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)

    df.loc[cond_origem, ["QTD", "SKU", "DESCRIÇÃO"]] = [0, "VAZIO", "VAZIO"]
    df = atualizar_endereco(df)
    return df, f"SKU {sku} movido de {origem_u} para {destino_u}.", enderecos_tocados


def _cmd_trocar(df: pd.DataFrame, sku1: str, endereco1: str, sku2: str, endereco2: str) -> tuple[pd.DataFrame, str, set[str]]:
    df = df.copy()
    end1_u, end2_u = _norm(endereco1), _norm(endereco2)
    sku1_u, sku2_u = _norm(sku1), _norm(sku2)
    enderecos_tocados: set[str] = {end1_u, end2_u}

    cond1 = _mask_sku_endereco(df, sku1_u, end1_u)
    cond2 = _mask_sku_endereco(df, sku2_u, end2_u)
    if not (cond1.any() and cond2.any()):
        return df, "Não foi possível encontrar ambos os SKUs e endereços.", enderecos_tocados

    idx1 = df.index[cond1][0]
    idx2 = df.index[cond2][0]

    fila1, torre1, nivel1 = (_norm(df.at[idx1, "FILA"]), _norm(df.at[idx1, "TORRE"]), _norm(df.at[idx1, "NÍVEL"]))
    fila2, torre2, nivel2 = (_norm(df.at[idx2, "FILA"]), _norm(df.at[idx2, "TORRE"]), _norm(df.at[idx2, "NÍVEL"]))

    df.loc[[idx1], ["FILA", "TORRE", "NÍVEL"]] = [fila2, torre2, nivel2]
    df.loc[[idx2], ["FILA", "TORRE", "NÍVEL"]] = [fila1, torre1, nivel1]

    df = atualizar_endereco(df)
    return df, f"SKUs {sku1} e {sku2} trocados entre {end1_u} e {end2_u}.", enderecos_tocados


def _cmd_tirar(df: pd.DataFrame, qtd: int, sku: str, endereco: str) -> tuple[pd.DataFrame, str, set[str]]:
    df = df.copy()
    end_u, sku_u = _norm(endereco), _norm(sku)
    enderecos_tocados: set[str] = {end_u}

    if qtd <= 0:
        return df, "Quantidade deve ser maior que zero.", enderecos_tocados

    cond = _mask_sku_endereco(df, sku_u, end_u)
    if not cond.any():
        return df, "SKU e endereço não encontrados.", enderecos_tocados

    df.loc[cond, "QTD"] = pd.to_numeric(df.loc[cond, "QTD"], errors="coerce").fillna(0).astype(int) - int(qtd)
    df.loc[cond, "QTD"] = df.loc[cond, "QTD"].clip(lower=0)

    df = aplicar_regra_zero_vazio(df)
    return df, f"{qtd} unidades retiradas do SKU {sku} no endereço {end_u}.", enderecos_tocados


def _cmd_add(df: pd.DataFrame, qtd: int, sku: str, endereco: str) -> tuple[pd.DataFrame, str, set[str]]:
    df = df.copy()
    end_u, sku_u = _norm(endereco), _norm(sku)
    enderecos_tocados: set[str] = {end_u}

    if qtd <= 0:
        return df, "Quantidade deve ser maior que zero.", enderecos_tocados

    cond_mesmo = _mask_sku_endereco(df, sku_u, end_u)
    end_norm = df["ENDEREÇO"].astype(str).str.upper().str.strip()
    sku_norm = df["SKU"].astype(str).str.upper().str.strip()
    cond_vazio_mesmo_end = (end_norm == end_u) & (sku_norm == "VAZIO")

    if cond_mesmo.any():
        df.loc[cond_mesmo, "QTD"] = (
            pd.to_numeric(df.loc[cond_mesmo, "QTD"], errors="coerce").fillna(0).astype(int) + int(qtd)
        )
        return df, f"{qtd} unidades adicionadas ao SKU {sku} no endereço {end_u}.", enderecos_tocados

    if cond_vazio_mesmo_end.any():
        desc_cond = sku_norm == sku_u
        if desc_cond.any():
            descricao = str(df.loc[desc_cond, "DESCRIÇÃO"].iloc[0])
            qtd_vazio = pd.to_numeric(df.loc[cond_vazio_mesmo_end, "QTD"], errors="coerce").fillna(0).astype(int).sum()
            df.loc[cond_vazio_mesmo_end, ["SKU", "DESCRIÇÃO", "QTD"]] = [sku_u, descricao, int(qtd_vazio + qtd)]
            return (
                df,
                f"SKU {sku} adicionado ao endereço {end_u} com {qtd} unidades (sobrescrevendo VAZIO).",
                enderecos_tocados,
            )
        else:
            return df, f"Descrição não encontrada para SKU {sku}.", enderecos_tocados

    sku_cond = sku_norm == sku_u
    if sku_cond.any():
        descricao = str(df.loc[sku_cond, "DESCRIÇÃO"].iloc[0])
        try:
            fila, torre, nivel = _split_endereco(end_u)
        except ValueError as e:
            return df, f"{e}", enderecos_tocados
        nova_linha = {
            "SKU": sku_u,
            "DESCRIÇÃO": descricao,
            "FILA": fila,
            "TORRE": torre,
            "NÍVEL": nivel,
            "ENDEREÇO": end_u,
            "QTD": int(qtd),
        }
        df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)
        return df, f"Novo SKU {sku} adicionado ao endereço {end_u} com {qtd} unidades.", enderecos_tocados

    return df, "SKU não encontrado para copiar a descrição.", enderecos_tocados


def _cmd_limpar(df: pd.DataFrame, endereco: str) -> tuple[pd.DataFrame, str, set[str]]:
    df = df.copy()
    end_u = _norm(endereco)
    enderecos_tocados: set[str] = {end_u}

    end_norm = df["ENDEREÇO"].fillna("").astype(str).str.upper().str.strip()
    df = df.loc[~end_norm.eq(end_u)].copy()

    try:
        fila, torre, nivel = _split_endereco(end_u)
    except ValueError as e:
        return df, f"{e}", enderecos_tocados

    nova_linha = {
        "SKU": "VAZIO",
        "DESCRIÇÃO": "VAZIO",
        "FILA": fila,
        "TORRE": torre,
        "NÍVEL": nivel,
        "ENDEREÇO": end_u,
        "QTD": 0,
    }
    df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)
    df = atualizar_endereco(df)
    return df, f"Endereço {end_u} limpo e marcado como VAZIO.", enderecos_tocados

# ============================
# Dispatcher
# ============================
def process_command(df: pd.DataFrame, comando: str) -> tuple[pd.DataFrame, str, set[str]]:
    """Processa comandos e retorna (df_atualizado, mensagem, enderecos_tocados)."""
    df = df.copy()
    comando_norm = (comando or "").strip()

    cols_min = ["SKU", "DESCRIÇÃO", "FILA", "TORRE", "NÍVEL", "ENDEREÇO", "QTD"]
    for c in (c for c in cols_min if c not in df.columns):
        df[c] = ""
    for col in ("SKU", "DESCRIÇÃO", "ENDEREÇO"):
        df[col] = df[col].astype(str)

    m = P_MUDAR.match(comando_norm)
    if m:
        return _cmd_mudar(df, m.group("sku"), m.group("origem"), m.group("destino"))

    m = P_TROCAR.match(comando_norm)
    if m:
        return _cmd_trocar(df, m.group("sku1"), m.group("end1"), m.group("sku2"), m.group("end2"))

    m = P_TIRAR.match(comando_norm)
    if m:
        return _cmd_tirar(df, int(m.group("qtd")), m.group("sku"), m.group("end"))

    m = P_ADD.match(comando_norm)
    if m:
        return _cmd_add(df, int(m.group("qtd")), m.group("sku"), m.group("end"))

    m = P_LIMPAR.match(comando_norm)
    if m:
        return _cmd_limpar(df, m.group("end"))

    return df, "Comando não reconhecido. Tente: mudar / trocar / tirar / add / limpar.", set()

# ============================
# Ordenação & salvar no Excel
# ============================
def sort_endereco_key(endereco: object) -> tuple[str, int, int]:
    if not isinstance(endereco, str):
        return ("ZZZ", 9999, 9999)
    parts = endereco.strip().upper().split("-")
    letra = parts[0] if len(parts) > 0 and parts[0] else "ZZZ"
    torre = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 9999
    nivel = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 9999
    return (letra, torre, nivel)


def reorganizar_e_salvar(df: pd.DataFrame, ws) -> pd.DataFrame:
    """Ordena por ENDEREÇO e escreve no worksheet, pulando a coluna 'ENDEREÇO'."""
    if df is None or df.empty:
        return df
    df_sorted = df.copy()
    df_sorted["__sort_key__"] = df_sorted["ENDEREÇO"].astype(str).apply(sort_endereco_key)
    df_sorted = df_sorted.sort_values("__sort_key__").drop(columns="__sort_key__")
    headers = [cell.value for cell in ws[1]]
    for i, row in enumerate(df_sorted.itertuples(index=False), start=2):
        row_dict = row._asdict() if hasattr(row, "_asdict") else dict(zip(df_sorted.columns, row))
        for col_idx, col_name in enumerate(headers, start=1):
            if col_name == "ENDEREÇO":
                continue
            value = row_dict.get(col_name, "")
            ws.cell(row=i, column=col_idx, value=value)
    return df_sorted

# ============================
# INTERFACE STREAMLIT
# ============================
st.title("MOVIMENTAR ESTOQUE")

with st.sidebar:
    uploaded_file = st.file_uploader("Faça upload da sua planilha (.xlsx)", type=["xlsx"])
    download_placeholder = st.empty()

    if st.button("Carregar do Banco de Dados"):
        st.session_state.df = _load_db_to_df()
        st.session_state.df_source = "db"
        if not st.session_state.df.empty:
            st.success("Dados do banco de dados carregados.")
        else:
            st.warning("Não foi possível carregar os dados do banco de dados.")

    if st.button("Registrar alterações no Banco de Dados"):
        if not st.session_state.df.empty:
            save_to_db(st.session_state.df)
        else:
            st.warning("Nada para salvar: DataFrame vazio.")

if uploaded_file is not None:
    st.session_state.df = _load_excel_to_df(uploaded_file)
    st.session_state.df_source = "upload"
    st.success("Planilha carregada.")

comando = st.text_input("Digite o comando:")
if st.button("Aplicar alteração"):
    if not st.session_state.df.empty:
        df_proc, msg, enderecos_tocados = process_command(st.session_state.df, comando)
        df_proc = _normalize_blanks(df_proc)
        df_proc = aplicar_regra_zero_vazio(df_proc)
        df_proc = drop_duplicates(df_proc)
        df_proc = regra_sku_vazio(df_proc, enderecos_afetados=enderecos_tocados)
        df_proc = atualizar_endereco(df_proc)

        st.session_state.df = df_proc
        st.session_state.comandos_acumulados.append((comando, msg))
        st.success(msg)
        st.dataframe(df_proc, use_container_width=True)
    else:
        st.warning("Nenhum dado foi carregado. Faça upload ou carregue do banco.")

if st.checkbox("Mostrar comandos acumulados"):
    st.write("Comandos acumulados:")
    for comando_txt, msg_txt in st.session_state.comandos_acumulados:
        st.write(f"- {comando_txt}: {msg_txt}")

# Download do Excel atualizado (quando origem for upload)
if uploaded_file and not st.session_state.df.empty:
    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active

        df_sorted = reorganizar_e_salvar(st.session_state.df, ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        with st.sidebar:
            download_placeholder.download_button(
                label="Baixar planilha atualizada",
                data=output.getvalue(),
                file_name="Estoque_atualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception as e:
        st.error(f"Não foi possível preparar o arquivo para download: {e}")

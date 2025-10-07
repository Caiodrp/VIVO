import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

# Função para carregar e consolidar as planilhas
def load_and_consolidate_data():
    # Caminho da pasta SQ no OneDrive
    folder_path = r'C:\Users\40418197\Desktop\robo_conferir\SQ'
    all_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]

    # Lista para armazenar os DataFrames
    data_frames = []

    for file in all_files:
        df = pd.read_excel(file)
        data_frames.append(df)

    # Concatenando todos os DataFrames em um único DataFrame
    consolidated_data = pd.concat(data_frames, ignore_index=True)

    # Removendo linhas com valores NaN
    consolidated_data = consolidated_data.dropna()

    # Removendo linhas duplicadas
    consolidated_data = consolidated_data.drop_duplicates()

    # Renomeando a coluna "Nº material" para "Descrição"
    consolidated_data.rename(columns={'Nº material': 'Descrição'}, inplace=True)

    # Convertendo colunas para os tipos apropriados
    consolidated_data['Fornecim.'] = consolidated_data['Fornecim.'].astype('int')
    consolidated_data['Nº de série'] = consolidated_data['Nº de série'].astype('int')
    consolidated_data['Material'] = consolidated_data['Material'].astype('object')
    consolidated_data['Descrição'] = consolidated_data['Descrição'].astype('object')

    return consolidated_data

# Função para atualizar a planilha "Fechamento Caixa"
def update_fechamento_caixa(fornecimento, volume_nota):
    file_path = r'C:\Users\40418197\Desktop\robo_conferir\Fechamento Caixa.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active

    # Encontrar a próxima linha vazia
    next_row = ws.max_row + 1

    # Encontrar os índices das colunas pelo nome
    col_indices = {cell.value: cell.column_letter for cell in ws[1]}

    # Preencher os dados na planilha
    ws[f"{col_indices['Fornecimento']}{next_row}"] = fornecimento
    ws[f"{col_indices['Volume Nota']}{next_row}"] = volume_nota
    ws[f"{col_indices['DVR']}{next_row}"] = 1  # DVR definido como 1
    ws[f"{col_indices['Hora Inicial']}{next_row}"] = datetime.now().strftime('%H:%M:%S')  # Hora Inicial
    ws[f"{col_indices['Hora Final']}{next_row}"] = datetime.now().strftime('%H:%M:%S')  # Hora Final (por enquanto igual à inicial)
    ws[f"{col_indices['Data']}{next_row}"] = datetime.now().strftime('%d/%m/%Y')  # Data

    # Salvar a planilha
    wb.save(file_path)

# Título da aplicação
st.title('Automação de Conferencia')

# Layout para os botões de carregar dados e finalizar conferência
col1, col2 = st.columns([1, 1])

with col1:
    if st.button('Carregar Dados'):
        st.session_state.data = load_and_consolidate_data()
        st.success('Dados carregados e consolidados com sucesso!')

# Verifica se os dados foram carregados
if 'data' in st.session_state:
    data = st.session_state.data

    with col2:
        if st.button('Finalizar Conferência'):
            fornecimento_input = st.session_state.get('fornecimento_input', None)
            volume_nota_input = st.session_state.get('volume_nota_input', None)
            if fornecimento_input and volume_nota_input:
                update_fechamento_caixa(fornecimento_input, volume_nota_input)
                st.success('Planilha "Fechamento Caixa" atualizada com sucesso!')

    # Widget para inserir o Fornecimento
    fornecimento_input = st.text_input('Insira o Fornecimento:')
    st.session_state.fornecimento_input = fornecimento_input

    if fornecimento_input:
        # Convertendo o fornecimento_input para string e depois para object
        try:
            fornecimento_input = int(fornecimento_input)
        except ValueError:
            st.write('Por favor, insira um número válido para o Fornecimento.')
            st.stop()

        # Filtrando os dados pelo Fornecimento inserido
        dados_filtrados = data[data['Fornecim.'] == fornecimento_input]
        
        # Mostrar número de volumes (coluna VOL)
        volumes = dados_filtrados['VOL'].unique()

        if len(volumes) == 1:
            st.warning(f"ATENÇÃO: {volumes[0]} VOLUMES")
        elif len(volumes) > 1:
            st.warning(f"ATENÇÃO: Múltiplos valores de VOL encontrados: {', '.join(map(str, volumes))}")
        else:
            st.warning("ATENÇÃO: Nenhum valor de VOL encontrado para este fornecimento.")

        if not dados_filtrados.empty:
            # # Mostrando o valor do volume correspondente ao fornecimento
            # volume_correspondente = dados_filtrados['Volume'].iloc[0]
            # st.write(f"**TOTAL DE VOLUMES: {volume_correspondente}**")

            # Widget para inserir o Volume Nota
            volume_nota_input = st.number_input('Insira o Volume Nota:', min_value=1)
            st.session_state.volume_nota_input = volume_nota_input

            # Widget para inserir seriais
            seriais_input = st.text_area('Insira os Seriais (um por linha):')

            if seriais_input:
                seriais_list = seriais_input.split('\n')
                seriais_list = [serial.strip() for serial in seriais_list if serial.strip()]

                # Removendo seriais duplicados inseridos pelo usuário
                seriais_list = list(set(seriais_list))

                # Verificando se a quantidade de seriais inseridos é igual ao total calculado
                total_qtd = len(dados_filtrados)
                if len(seriais_list) != total_qtd:
                    st.error(f"A quantidade de seriais inseridos ({len(seriais_list)}) não corresponde ao total calculado ({total_qtd}).")
                    if len(seriais_list) > total_qtd:
                        st.write("Seriais a mais:")
                        seriais_a_mais = set(seriais_list) - set(dados_filtrados['Nº de série'].astype(str))
                        for serial in seriais_a_mais:
                            st.write(f"Serial: {serial}")
                    else:
                        st.write("Seriais faltando:")
                        seriais_faltando = set(dados_filtrados['Nº de série'].astype(str)) - set(seriais_list)
                        if seriais_faltando:
                            dados_faltantes = dados_filtrados[dados_filtrados['Nº de série'].astype(str).isin(seriais_faltando)]
                            st.write('Dados dos seriais faltantes:')
                            st.dataframe(dados_faltantes[['Material', 'Nº de série', 'Descrição']].reset_index(drop=True))
                else:
                    try:
                        seriais_list = [int(serial) for serial in seriais_list]
                    except ValueError:
                        st.write('Por favor, insira apenas números válidos para os Seriais.')
                        st.stop()

                    seriais_presentes = dados_filtrados[dados_filtrados['Nº de série'].isin(seriais_list)]

                    if len(seriais_list) == len(seriais_presentes):
                        st.success('Todos os seriais estão presentes.')
                    else:
                        st.write('Os seguintes seriais estão faltando:')
                        seriais_faltantes = set(dados_filtrados['Nº de série']) - set(seriais_list)
                        if seriais_faltantes:
                            dados_faltantes = dados_filtrados[dados_filtrados['Nº de série'].isin(seriais_faltantes)]
                            st.write('Dados dos seriais faltantes:')
                            st.dataframe(dados_faltantes[['Material', 'Nº de série', 'Descrição']].reset_index(drop=True), use_container_width=True)

            # Mostrando a tabela filtrada com colunas específicas
            st.write('Dados Filtrados:')
            st.dataframe(dados_filtrados[['Material', 'Nº de série', 'Descrição']].reset_index(drop=True))

        else:
            st.write('Fornecimento não encontrado.')